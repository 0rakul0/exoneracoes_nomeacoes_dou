from __future__ import annotations

import re
import unicodedata
from datetime import datetime
from pathlib import Path

import pandas as pd


# ============================================================
# TRATAMENTO — O QUE MUDA / ONDE MUDA
# ============================================================
#
# Entradas:
#   saida/ISP/circunscricao/ISP_circunscricao_mudancas.csv
#   \\10.9.11.20\...\cria_altera_anual.xlsx
#
# Saídas:
#   saida/ISP/o_que_muda_onde_muda/ISP_o_que_muda_onde_muda.csv
#   saida/ISP/o_que_muda_onde_muda/ISP_o_que_muda_onde_muda_resumo.csv
#   saida/ISP/o_que_muda_onde_muda/ISP_o_que_muda_onde_muda.xlsx
#
# A planilha anual é uma matriz:
#   linhas = unidades numéricas alteradas
#   colunas = datas dos atos
#   células = ação (Altera, Implanta, Extingue...)
#
# Ela é usada como referência auxiliar de "onde muda" para unidades
# extraídas do texto, especialmente AISP/CISP.
# O texto das normativas continua sendo a fonte principal.
# ============================================================


BASE_DIR = Path(__file__).resolve().parent.parent.parent

ARQUIVO_MUDANCAS = BASE_DIR / "saida" / "ISP" / "circunscricao" / "ISP_circunscricao_mudancas.csv"
PASTA_SAIDA = BASE_DIR / "saida" / "ISP" / "o_que_muda_onde_muda"

PLANILHA_ANUAL = Path(
    r"\\10.9.11.20\Projetos\01- Coordenadoria\Coordenadoria Projetos\03 - Dados\08 - Áreas de segurança\Limites RISP-AISP-CISP\cria_altera_anual.xlsx"
)
PLANILHA_ANUAL_FALLBACK = BASE_DIR / ".cache" / "cria_altera_anual.xlsx"

GERAR_EXCEL = True

MESES = {
    "JANEIRO": 1,
    "FEVEREIRO": 2,
    "MARCO": 3,
    "MARÇO": 3,
    "ABRIL": 4,
    "MAIO": 5,
    "JUNHO": 6,
    "JULHO": 7,
    "AGOSTO": 8,
    "SETEMBRO": 9,
    "OUTUBRO": 10,
    "NOVEMBRO": 11,
    "DEZEMBRO": 12,
}

LIMITES_UNIDADE = {
    "CISP": 200,
    "AISP": 50,
    "RISP": 10,
    "BPM": 100,
}


def limpar_texto(value: object) -> str:
    if pd.isna(value):
        return ""

    return re.sub(r"\s+", " ", str(value)).strip()


def normalizar_texto(value: object) -> str:
    text = limpar_texto(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = text.upper()
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def normalizar_numero(value: object) -> str:
    text = limpar_texto(value)
    match = re.search(r"\d+", text)
    return str(int(match.group(0))) if match else ""


def juntar_unicos(values: list[object]) -> str:
    saida = []
    vistos = set()

    for value in values:
        item = limpar_texto(value)
        if not item:
            continue

        chave = normalizar_texto(item)
        if chave in vistos:
            continue

        vistos.add(chave)
        saida.append(item)

    return " | ".join(saida)


def resolver_planilha_anual() -> Path | None:
    if PLANILHA_ANUAL.exists():
        return PLANILHA_ANUAL

    if PLANILHA_ANUAL_FALLBACK.exists():
        return PLANILHA_ANUAL_FALLBACK

    return None


def parse_data_norma_texto(value: object) -> pd.Timestamp | pd.NaT:
    texto = normalizar_texto(value)
    match = re.search(
        r"(\d{1,2})\s+DE\s+([A-ZÇ]+)\s+DE\s+(\d{4})",
        texto,
        flags=re.I,
    )

    if not match:
        return pd.NaT

    dia = int(match.group(1))
    mes = MESES.get(match.group(2).upper())
    ano = int(match.group(3))

    if not mes:
        return pd.NaT

    return pd.Timestamp(datetime(ano, mes, dia).date())


def carregar_mudancas(path: Path) -> pd.DataFrame:
    df = pd.read_csv(path, dtype=str, keep_default_na=False)
    df["data_publicacao_dt"] = pd.to_datetime(df["data_publicacao"], errors="coerce")
    df["data_norma_dt"] = df["data_norma_texto"].apply(parse_data_norma_texto)
    df["texto_analise_onde"] = df.apply(montar_texto_analise, axis=1)
    return df


def montar_texto_analise(row: pd.Series) -> str:
    campos = [
        row.get("norma_id", ""),
        row.get("ementa_limpa", ""),
        row.get("trechos_circunscricao", ""),
        row.get("o_que_mudou_tratado", ""),
        row.get("texto_norma", ""),
    ]
    return limpar_texto(" ".join(str(campo or "") for campo in campos))


def carregar_referencia_anual(path: Path | None) -> tuple[pd.DataFrame, pd.DataFrame]:
    if path is None:
        return pd.DataFrame(), pd.DataFrame()

    plan1 = pd.read_excel(path, sheet_name="Plan1", dtype=str, keep_default_na=False)
    plan2 = pd.read_excel(path, sheet_name="Plan2", dtype=str, keep_default_na=False)

    coluna_unidade = plan1.columns[0]
    eventos = plan1.melt(
        id_vars=[coluna_unidade],
        var_name="data_evento",
        value_name="acao_planilha",
    )

    eventos = eventos[eventos["acao_planilha"].astype(str).str.strip().ne("")].copy()
    eventos = eventos.rename(columns={coluna_unidade: "unidade_planilha"})
    eventos["unidade_planilha"] = eventos["unidade_planilha"].apply(normalizar_numero)
    eventos["data_evento"] = pd.to_datetime(eventos["data_evento"], errors="coerce")
    eventos["acao_planilha"] = eventos["acao_planilha"].apply(limpar_texto)
    eventos = eventos[eventos["unidade_planilha"].ne("") & eventos["data_evento"].notna()].copy()

    coluna_unidade_plan2 = plan2.columns[0]
    obs = plan2.rename(columns={coluna_unidade_plan2: "unidade_planilha"}).copy()
    obs["unidade_planilha"] = obs["unidade_planilha"].apply(normalizar_numero)

    return eventos, obs


def numero_no_intervalo(tipo: str, numero: str) -> bool:
    limite = LIMITES_UNIDADE.get(tipo)
    if limite is None:
        return True

    try:
        valor = int(numero)
    except ValueError:
        return False

    return 1 <= valor <= limite


def numeros_apos_rotulo(texto: str, rotulo_pattern: str, tipo: str) -> list[str]:
    encontrados = []
    vistos = set()

    for match in re.finditer(rotulo_pattern, texto, flags=re.I):
        trecho = texto[match.end() : match.end() + 180]
        trecho = re.split(
            r"\.|;|\n| Art\.?\s*\d+|"
            r"\b(?:CISP|AISP|RISP|BPM)\b|"
            r"\b(?:CIRCUNSCRI[ÇC][ÃA]O(?:ES)?|[ÁA]REA(?:S)?|REGI[ÃA]O(?:ES)?)\s+INTEGRADA(?:S)?\s+DE\s+SEGURAN[ÇC]A\s+P[ÚU]BLICA\b|"
            r"\bCONSTANTES?\b|\bNA\s+RESOLU[ÇC][ÃA]O\b|\bRESOLU[ÇC][ÃA]O\b|\bDECRETO\b|\bLEI\b",
            trecho,
            maxsplit=1,
            flags=re.I,
        )[0]
        trecho = re.sub(
            r"\b(?:RESOLU[ÇC][ÃA]O|DECRETO|LEI|PORTARIA)\s+[A-Z]{0,12}\s*(?:N[º°O]\.?|Nº|N\.)?\s*\d{1,5}",
            "",
            trecho,
            flags=re.I,
        )

        for numero in re.findall(r"\b\d{1,3}\b", trecho):
            numero_norm = str(int(numero))
            if not numero_no_intervalo(tipo, numero_norm):
                continue
            if numero_norm not in vistos:
                vistos.add(numero_norm)
                encontrados.append(numero_norm)

    return encontrados


def extrair_unidades_numericas(texto: str) -> dict[str, list[str]]:
    unidades = {
        "CISP": numeros_apos_rotulo(
            texto,
            r"\b(?:CISP|CIRCUNSCRI[ÇC][ÃA]O(?:ES)?\s+INTEGRADA(?:S)?\s+DE\s+SEGURAN[ÇC]A\s+P[ÚU]BLICA)\b",
            "CISP",
        ),
        "AISP": numeros_apos_rotulo(
            texto,
            r"\b(?:AISP|[ÁA]REA(?:S)?\s+INTEGRADA(?:S)?\s+DE\s+SEGURAN[ÇC]A\s+P[ÚU]BLICA)\b",
            "AISP",
        ),
        "RISP": numeros_apos_rotulo(
            texto,
            r"\b(?:RISP|REGI[ÃA]O(?:ES)?\s+INTEGRADA(?:S)?\s+DE\s+SEGURAN[ÇC]A\s+P[ÚU]BLICA)\b",
            "RISP",
        ),
        "BPM": [],
    }

    cisp_por_dp = extrair_numeros_delegacias(texto)
    bpm = re.findall(r"\b(\d{1,3})\s*(?:[º°]|O)?\s*BPM\b", texto, flags=re.I)

    for numero in cisp_por_dp:
        numero_norm = str(int(numero))
        if not numero_no_intervalo("CISP", numero_norm):
            continue
        if numero_norm not in unidades["CISP"]:
            unidades["CISP"].append(numero_norm)

    for numero in bpm:
        numero_norm = str(int(numero))
        if not numero_no_intervalo("BPM", numero_norm):
            continue
        if numero_norm not in unidades["BPM"]:
            unidades["BPM"].append(numero_norm)

    return unidades


def extrair_numeros_delegacias(texto: str) -> list[str]:
    """
    Captura números de DP/CISP em formas como:
      26ª DP
      26ª. Delegacia Policial
      26ª e 44ª Delegacia Policial

    A última forma é comum em ementas e o número anterior ao "e"
    não fica imediatamente antes da palavra Delegacia.
    """
    encontrados = []
    vistos = set()

    patterns = [
        r"\b((?:\d{1,3}\s*(?:[ªº°]|A|O)?\.?\s*(?:,|E|e)?\s*){1,6})"
        r"(?:DP|DELEGACIA(?:\s+POLICIAL)?|DELEGACIA\s+DE\s+POL[ÍI]CIA)\b",
        r"\b(\d{1,3})\s*(?:[ªº°]|A|O)?\.?\s*DP\b",
    ]

    for pattern in patterns:
        for match in re.finditer(pattern, texto, flags=re.I):
            for numero in re.findall(r"\d{1,3}", match.group(1)):
                numero_norm = str(int(numero))
                if not numero_no_intervalo("CISP", numero_norm):
                    continue
                if numero_norm in vistos:
                    continue

                vistos.add(numero_norm)
                encontrados.append(numero_norm)

    return encontrados


def extrair_locais_nomeados(texto: str) -> dict[str, list[str]]:
    locais = {"MUNICIPIO": [], "BAIRRO": []}
    patterns = {
        "MUNICIPIO": r"\bMunic[íi]pio(?:s)?\s+de\s+([^.;|]+)",
        "BAIRRO": r"\bBairro(?:s)?\s+de\s+([^.;|]+)",
    }

    for tipo, pattern in patterns.items():
        vistos = set()
        for match in re.finditer(pattern, texto, flags=re.I):
            trecho = limpar_texto(match.group(1))
            trecho = re.split(
                r"\bcuja\b|\bconforme\b|\bna forma\b|\bpassa\b|\be\s+d[áa]\b|\breorganizando\b|\bat[ée]\b|\bsob\b|\bque\b",
                trecho,
                maxsplit=1,
                flags=re.I,
            )[0]

            for parte in re.split(r",|\se\s", trecho):
                item = limpar_texto(parte)
                item = re.sub(r"^(?:o|a|os|as|de|do|da|dos|das)\s+", "", item, flags=re.I)
                item = re.sub(r"\s*\([^)]*$", "", item).strip()
                item = item[:80].strip()
                item_norm = normalizar_texto(item)
                if re.match(r"^(PERMANECERAO|PERMANECERA|TENDO|COMO|CUJO|CUJA|CONFORME|ATE|ATUAL)$", item_norm):
                    continue
                if not item or len(item) < 3:
                    continue

                chave = item_norm
                if chave in vistos:
                    continue

                vistos.add(chave)
                locais[tipo].append(item)

    return locais


def acao_principal(tipo_mudanca: str, acao_planilha: str = "") -> str:
    acao_planilha_norm = normalizar_texto(acao_planilha)
    if acao_planilha_norm:
        mapa_planilha = {
            "ALTERA": "alteração",
            "ALTERA*": "alteração",
            "IMPLANTA": "criação/instituição",
            "CRIA": "criação/instituição",
            "EXTINGUE": "extinção",
            "DESATIVA": "extinção/desativação",
            "REATIVA": "reativação",
        }
        return mapa_planilha.get(acao_planilha_norm, acao_planilha)

    texto = normalizar_texto(tipo_mudanca)

    prioridades = [
        ("revogação", "REVOGACAO"),
        ("extinção", "EXTINCAO"),
        ("criação/instituição", "CRIACAO/INSTITUICAO"),
        ("alteração", "ALTERACAO"),
        ("competência/responsabilidade", "COMPETENCIA/RESPONSABILIDADE"),
        ("vigência", "VIGENCIA"),
    ]

    for saida, termo in prioridades:
        if termo in texto:
            return saida

    return "não classificado"


def buscar_eventos_planilha(
    eventos: pd.DataFrame,
    unidade_numero: str,
    data_norma: pd.Timestamp | pd.NaT,
    data_publicacao: pd.Timestamp | pd.NaT,
) -> tuple[str, str, str]:
    if eventos.empty or not unidade_numero:
        return "", "", ""

    subset = eventos[eventos["unidade_planilha"].eq(unidade_numero)].copy()
    if subset.empty:
        return "", "", ""

    candidatos = []

    for label, data_ref in [("data_norma", data_norma), ("data_publicacao", data_publicacao)]:
        if pd.isna(data_ref):
            continue

        exatos = subset[subset["data_evento"].dt.date.eq(data_ref.date())]
        if not exatos.empty:
            candidatos.append((label, exatos))

    if not candidatos:
        datas_ref = [data for data in [data_norma, data_publicacao] if not pd.isna(data)]
        for data_ref in datas_ref:
            subset["distancia_dias"] = (subset["data_evento"] - data_ref).abs().dt.days
            proximos = subset[subset["distancia_dias"].le(10)]
            if not proximos.empty:
                candidatos.append(("proxima_ate_10_dias", proximos.sort_values("distancia_dias")))
                break

    if not candidatos:
        return "", "", ""

    tipo_match, rows = candidatos[0]
    rows = rows.copy()
    rows["data_evento_str"] = rows["data_evento"].dt.strftime("%Y-%m-%d")

    return (
        juntar_unicos(rows["acao_planilha"].tolist()),
        juntar_unicos(rows["data_evento_str"].tolist()),
        tipo_match,
    )


def montar_linhas_onde_muda(
    df_mudancas: pd.DataFrame,
    eventos_planilha: pd.DataFrame,
    obs_planilha: pd.DataFrame,
) -> pd.DataFrame:
    linhas = []

    obs_by_unidade = {}
    if not obs_planilha.empty:
        for _, row in obs_planilha.iterrows():
            unidade_planilha = row.get("unidade_planilha", "")
            if not unidade_planilha:
                continue
            obs_by_unidade[unidade_planilha] = {
                "nao_achou_documentacao": limpar_texto(row.get("NÃO ACHOU DOCUMENTAÇÃO", "")),
                "verificada": limpar_texto(row.get("VERIFICADA", "")),
                "area_correta": limpar_texto(row.get("ESTÁ COM ÁREA CORRETA", "")),
                "obs_planilha": limpar_texto(row.get("OBS", "")),
            }

    for _, row in df_mudancas.iterrows():
        texto = row.get("texto_analise_onde", "")
        unidades = extrair_unidades_numericas(texto)
        locais = extrair_locais_nomeados(texto)

        unidades_encontradas = []
        for tipo, numeros in unidades.items():
            for numero in numeros:
                unidades_encontradas.append((tipo, numero, ""))

        for tipo, nomes in locais.items():
            for nome in nomes:
                unidades_encontradas.append((tipo, "", nome))

        if not unidades_encontradas:
            unidades_encontradas.append(("NAO_IDENTIFICADO", "", ""))

        for unidade_tipo, unidade_numero, unidade_nome in unidades_encontradas:
            acao_planilha = ""
            data_evento_planilha = ""
            match_planilha = ""
            obs = {}

            if unidade_tipo in {"AISP", "CISP"}:
                acao_planilha, data_evento_planilha, match_planilha = buscar_eventos_planilha(
                    eventos=eventos_planilha,
                    unidade_numero=unidade_numero,
                    data_norma=row.get("data_norma_dt", pd.NaT),
                    data_publicacao=row.get("data_publicacao_dt", pd.NaT),
                )
                obs = obs_by_unidade.get(unidade_numero, {})

            linhas.append(
                {
                    "ano_publicacao": row.get("ano_publicacao", ""),
                    "data_publicacao": row.get("data_publicacao", ""),
                    "data_norma": "" if pd.isna(row.get("data_norma_dt", pd.NaT)) else row.get("data_norma_dt").strftime("%Y-%m-%d"),
                    "norma_id": row.get("norma_id", ""),
                    "link": row.get("link", ""),
                    "unidade_tipo": unidade_tipo,
                    "unidade_numero": unidade_numero,
                    "unidade_nome": unidade_nome,
                    "acao_detectada": acao_principal(row.get("tipo_mudanca_detectada", ""), acao_planilha),
                    "tipo_mudanca_detectada": row.get("tipo_mudanca_detectada", ""),
                    "acao_planilha": acao_planilha,
                    "data_evento_planilha": data_evento_planilha,
                    "match_planilha": match_planilha,
                    "nao_achou_documentacao": obs.get("nao_achou_documentacao", ""),
                    "verificada": obs.get("verificada", ""),
                    "area_correta": obs.get("area_correta", ""),
                    "obs_planilha": obs.get("obs_planilha", ""),
                    "ementa_limpa": row.get("ementa_limpa", ""),
                    "o_que_mudou": row.get("o_que_mudou_tratado", ""),
                    "trechos_circunscricao": row.get("trechos_circunscricao", ""),
                    "arquivo_markdown": row.get("arquivo_markdown", ""),
                }
            )

    return pd.DataFrame(linhas)


def montar_resumo(df_onde: pd.DataFrame, eventos_planilha: pd.DataFrame) -> pd.DataFrame:
    normas_distintas = 0
    if not df_onde.empty:
        normas_distintas = len(
            df_onde[["data_publicacao", "data_norma", "norma_id", "link"]].drop_duplicates()
        )

    linhas = [
        ("linhas_onde_muda", len(df_onde)),
        ("eventos_planilha_anual", len(eventos_planilha)),
        ("normas_distintas", normas_distintas),
        ("linhas_com_match_planilha", int(df_onde["match_planilha"].astype(str).str.strip().ne("").sum()) if not df_onde.empty else 0),
        ("linhas_sem_unidade_identificada", int(df_onde["unidade_tipo"].eq("NAO_IDENTIFICADO").sum()) if not df_onde.empty else 0),
    ]

    if not df_onde.empty:
        for tipo, qtd in df_onde["unidade_tipo"].value_counts().items():
            linhas.append((f"unidade_tipo: {tipo}", int(qtd)))

        for acao, qtd in df_onde["acao_detectada"].value_counts().items():
            linhas.append((f"acao_detectada: {acao}", int(qtd)))

    return pd.DataFrame(linhas, columns=["indicador", "valor"])


def montar_resumo_por_norma(df_onde: pd.DataFrame) -> pd.DataFrame:
    if df_onde.empty:
        return pd.DataFrame()

    def filtrar_tipo(group: pd.DataFrame, tipo: str, coluna: str) -> str:
        return juntar_unicos(group.loc[group["unidade_tipo"].eq(tipo), coluna].tolist())

    linhas = []
    chaves = ["data_publicacao", "data_norma", "norma_id", "link"]

    for chave, group in df_onde.groupby(chaves, dropna=False, sort=True):
        row = group.iloc[0]
        linhas.append(
            {
                "data_publicacao": chave[0],
                "data_norma": chave[1],
                "norma_id": chave[2],
                "link": chave[3],
                "acoes_detectadas": juntar_unicos(group["acao_detectada"].tolist()),
                "tipo_mudanca_detectada": row.get("tipo_mudanca_detectada", ""),
                "cisp": filtrar_tipo(group, "CISP", "unidade_numero"),
                "aisp": filtrar_tipo(group, "AISP", "unidade_numero"),
                "risp": filtrar_tipo(group, "RISP", "unidade_numero"),
                "bpm": filtrar_tipo(group, "BPM", "unidade_numero"),
                "municipios": filtrar_tipo(group, "MUNICIPIO", "unidade_nome"),
                "bairros": filtrar_tipo(group, "BAIRRO", "unidade_nome"),
                "matches_planilha": juntar_unicos(group["match_planilha"].tolist()),
                "acoes_planilha": juntar_unicos(group["acao_planilha"].tolist()),
                "datas_evento_planilha": juntar_unicos(group["data_evento_planilha"].tolist()),
                "ementa_limpa": row.get("ementa_limpa", ""),
                "o_que_mudou": row.get("o_que_mudou", ""),
                "trechos_circunscricao": row.get("trechos_circunscricao", ""),
                "arquivo_markdown": row.get("arquivo_markdown", ""),
            }
        )

    return pd.DataFrame(linhas)


def formatar_celula_matriz(group: pd.DataFrame) -> str:
    partes = []
    vistos = set()

    for _, row in group.iterrows():
        acao = limpar_texto(row.get("acao_planilha", "")) or limpar_texto(row.get("acao_detectada", ""))
        norma = limpar_texto(row.get("norma_id", ""))
        data_publicacao = pd.to_datetime(row.get("data_publicacao", ""), errors="coerce")
        data_diario = "" if pd.isna(data_publicacao) else f"D.O. {data_publicacao:%Y-%m-%d}"

        if not acao and not norma:
            continue

        item = acao
        if norma:
            item = f"{acao} - {norma}" if acao else norma
        if data_diario and norma:
            item = f"{item} [{data_diario}]"

        chave = normalizar_texto(item)
        if chave in vistos:
            continue

        vistos.add(chave)
        partes.append(item)

    return " | ".join(partes)


def montar_matriz_cisp(df_onde: pd.DataFrame) -> pd.DataFrame:
    cisp = df_onde[
        df_onde["unidade_tipo"].eq("CISP")
        & df_onde["unidade_numero"].astype(str).str.strip().ne("")
        & df_onde["data_norma"].astype(str).str.strip().ne("")
    ].copy()

    if cisp.empty:
        return pd.DataFrame(columns=["CISP"])

    cisp["cisp_num"] = cisp["unidade_numero"].astype(int)
    cisp["data_coluna"] = pd.to_datetime(cisp["data_norma"], errors="coerce")
    cisp = cisp[cisp["data_coluna"].notna()].copy()

    linhas = []
    for (cisp_num, data_coluna), group in cisp.groupby(["cisp_num", "data_coluna"], sort=True):
        linhas.append(
            {
                "CISP": cisp_num,
                "data_coluna": data_coluna,
                "valor": formatar_celula_matriz(group),
            }
        )

    base = pd.DataFrame(linhas)
    matriz = base.pivot_table(
        index="CISP",
        columns="data_coluna",
        values="valor",
        aggfunc=lambda values: juntar_unicos(list(values)),
        fill_value="",
    )

    matriz = matriz.sort_index()
    matriz.columns = [pd.Timestamp(col).strftime("%Y-%m-%d") for col in matriz.columns]
    matriz = matriz.reset_index()
    matriz["CISP"] = matriz["CISP"].astype(int)

    return matriz


def ajustar_excel(path: Path) -> None:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = load_workbook(path)
    cores = {
        "Matriz_CISP": "1F4E78",
        "O_que_muda_onde": "305496",
        "Resumo": "1F4E78",
        "Resumo_por_norma": "1F4E78",
        "Planilha_anual": "548235",
        "Dicionario": "7030A0",
    }

    larguras = {
        "ano_publicacao": 12,
        "CISP": 10,
        "data_publicacao": 14,
        "data_norma": 14,
        "norma_id": 28,
        "link": 45,
        "unidade_tipo": 18,
        "unidade_numero": 16,
        "unidade_nome": 30,
        "acoes_detectadas": 32,
        "cisp": 26,
        "aisp": 24,
        "risp": 20,
        "bpm": 24,
        "municipios": 45,
        "bairros": 45,
        "matches_planilha": 24,
        "acoes_planilha": 24,
        "datas_evento_planilha": 28,
        "acao_detectada": 28,
        "tipo_mudanca_detectada": 36,
        "acao_planilha": 18,
        "data_evento_planilha": 22,
        "match_planilha": 20,
        "obs_planilha": 55,
        "ementa_limpa": 55,
        "o_que_mudou": 70,
        "trechos_circunscricao": 65,
        "arquivo_markdown": 42,
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        cor = cores.get(sheet_name, "1F4E78")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=cor)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        header_by_col = {idx: str(cell.value or "") for idx, cell in enumerate(ws[1], start=1)}

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if header_by_col.get(cell.column) == "link" and cell.row > 1 and cell.value:
                    cell.hyperlink = str(cell.value)
                    cell.style = "Hyperlink"

        for idx, header in header_by_col.items():
            ws.column_dimensions[get_column_letter(idx)].width = larguras.get(header, 18)

        for row_idx in range(2, min(ws.max_row, 300) + 1):
            ws.row_dimensions[row_idx].height = 55

    wb.save(path)


def exportar(
    df_onde: pd.DataFrame,
    df_resumo: pd.DataFrame,
    df_resumo_norma: pd.DataFrame,
    df_matriz_cisp: pd.DataFrame,
    eventos_planilha: pd.DataFrame,
) -> None:
    PASTA_SAIDA.mkdir(parents=True, exist_ok=True)

    xlsx_path = PASTA_SAIDA / "ISP_o_que_muda_onde_muda.xlsx"

    if GERAR_EXCEL:
        dicionario = pd.DataFrame(
            [
                ("unidade_tipo", "Tipo de lugar/unidade extraído do texto: CISP, AISP, RISP, BPM, MUNICIPIO, BAIRRO ou NAO_IDENTIFICADO."),
                ("unidade_numero", "Número da unidade quando houver, por exemplo CISP 42 ou AISP 9."),
                ("unidade_nome", "Nome textual quando o local for município ou bairro."),
                ("acao_detectada", "Ação principal, priorizando a ação da planilha anual quando há match de CISP/data."),
                ("acao_planilha", "Ação registrada em cria_altera_anual.xlsx para a unidade numérica/data."),
                ("match_planilha", "Tipo de casamento da unidade numérica com a matriz anual: data_norma, data_publicacao ou proxima_ate_10_dias."),
                ("o_que_mudou", "Resumo extraído dos artigos/ementa da normativa."),
            ],
            columns=["campo", "descricao"],
        )

        xlsx_export_path = xlsx_path
        try:
            writer_context = pd.ExcelWriter(xlsx_export_path, engine="openpyxl")
        except PermissionError:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            xlsx_export_path = PASTA_SAIDA / f"ISP_o_que_muda_onde_muda_{timestamp}.xlsx"
            print(f"\nAVISO: XLSX principal bloqueado; salvando cópia em {xlsx_export_path}")
            writer_context = pd.ExcelWriter(xlsx_export_path, engine="openpyxl")

        with writer_context as writer:
            df_matriz_cisp.to_excel(writer, sheet_name="Matriz_CISP", index=False)
            df_resumo_norma.to_excel(writer, sheet_name="Resumo_por_norma", index=False)
            df_onde.to_excel(writer, sheet_name="O_que_muda_onde", index=False)
            df_resumo.to_excel(writer, sheet_name="Resumo", index=False)
            eventos_planilha.to_excel(writer, sheet_name="Planilha_anual", index=False)
            dicionario.to_excel(writer, sheet_name="Dicionario", index=False)

        ajustar_excel(xlsx_export_path)

    print("\nArquivos gerados:")
    if GERAR_EXCEL:
        print(f"- {xlsx_export_path}")


def main() -> None:
    print("=" * 70)
    print("TRATAMENTO — O QUE MUDA / ONDE MUDA")
    print("=" * 70)

    if not ARQUIVO_MUDANCAS.exists():
        print("\nERRO: arquivo de mudanças não encontrado.")
        print(ARQUIVO_MUDANCAS)
        raise SystemExit(1)

    planilha_path = resolver_planilha_anual()
    if planilha_path is None:
        print("\nAVISO: planilha anual não encontrada; gerando apenas extração textual.")
    else:
        print(f"\nPlanilha anual: {planilha_path}")

    print(f"Arquivo de mudanças: {ARQUIVO_MUDANCAS}")
    print(f"Pasta de saída: {PASTA_SAIDA}")

    df_mudancas = carregar_mudancas(ARQUIVO_MUDANCAS)
    eventos_planilha, obs_planilha = carregar_referencia_anual(planilha_path)
    df_onde = montar_linhas_onde_muda(df_mudancas, eventos_planilha, obs_planilha)
    df_resumo_norma = montar_resumo_por_norma(df_onde)
    df_matriz_cisp = montar_matriz_cisp(df_onde)
    df_resumo = montar_resumo(df_onde, eventos_planilha)

    exportar(df_onde, df_resumo, df_resumo_norma, df_matriz_cisp, eventos_planilha)

    print("\nResumo:")
    print(df_resumo.to_string(index=False))
    print("\nProcesso concluído com sucesso.")


if __name__ == "__main__":
    main()
