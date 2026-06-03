from __future__ import annotations

import json
import re
import unicodedata
from pathlib import Path

import pandas as pd


# ============================================================
# TRATAMENTO — CIRCUNSCRIÇÃO / ISP
# XLSX LIMPO COM COLUNA link
# ============================================================
#
# Entrada:
#   saida/ISP/DOERJ_normativas_circunscricao.csv
#
# Saídas:
#   saida/ISP/circunscricao/ISP_circunscricao_tratado.csv
#   saida/ISP/circunscricao/ISP_circunscricao_mudancas.csv
#   saida/ISP/circunscricao/ISP_circunscricao_resumo.csv
#   saida/ISP/circunscricao/ISP_circunscricao_tratado.xlsx
#
# O Excel terá uma coluna chamada:
#   link
#
# Essa coluna vem diretamente do extrator.
# ============================================================


BASE_DIR = Path(__file__).resolve().parent.parent.parent

ARQUIVO_ENTRADA = BASE_DIR / "saida" / "ISP" / "DOERJ_normativas_circunscricao.csv"
PASTA_SAIDA = BASE_DIR / "saida" / "ISP" / "circunscricao"

GERAR_EXCEL = True


def normalizar_texto(value: object) -> str:
    if pd.isna(value):
        return ""

    text = str(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = text.upper()
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def limpar_texto(value: object) -> str:
    if pd.isna(value):
        return ""

    return re.sub(r"\s+", " ", str(value)).strip()


def garantir_colunas(df: pd.DataFrame, colunas: list[str]) -> pd.DataFrame:
    df = df.copy()

    for coluna in colunas:
        if coluna not in df.columns:
            df[coluna] = ""

    return df


CIRCUNSCRICAO_RE = re.compile(
    r"\bCIRCUNSCRI(?:CAO|COES|CIONAL|CIONAIS|CIONARIA|CIONARIAS|CIONADO|CIONADOS|CIONADA|CIONADAS)\b",
    flags=re.I,
)

AREA_ATUACAO_RE = re.compile(
    r"\b(?:AISP|RISP|CISP)\b|"
    r"\b(?:AREA|AREAS)\s+(?:INTEGRADA|INTEGRADAS)\s+DE\s+SEGURANCA\s+PUBLICA\b|"
    r"\b(?:REGIAO|REGIOES)\s+(?:INTEGRADA|INTEGRADAS)\s+DE\s+SEGURANCA\s+PUBLICA\b|"
    r"\b(?:CIRCUNSCRICAO|CIRCUNSCRICOES)\s+(?:INTEGRADA|INTEGRADAS)\s+DE\s+SEGURANCA\s+PUBLICA\b|"
    r"\b(?:AREA|AREAS)\s+DE\s+(?:ATUACAO|POLICIAMENTO)\b|"
    r"\bLIMITES?\s+(?:TERRITORIAIS|CIRCUNSCRICIONAIS)\b|"
    r"\b(?:ALTERA|ALTERAR|MODIFICA|MODIFICAR|RESTAURA|RESTAURAR|IMPLANTA|IMPLANTAR|CRIA|CRIAR|EXTINGUE|EXTINGUIR|REVOGA|REVOGAR|TRANSFERE|TRANSFERIR|TRANSFERINDO)\b.{0,160}\b(?:DP|DELEGACIA|BPM|AISP|RISP|CISP)\b|"
    r"\b(?:PASSA\s+A\s+INTEGRAR|PASSANDO\s+A|PASSA\s+A\s+VIGORAR)\b.{0,160}\b(?:DP|DELEGACIA|BPM|AISP|RISP|CISP)\b|"
    r"\b(?:DP|DELEGACIA|BPM|AISP|RISP|CISP)\b.{0,160}\b(?:AREA|LIMITES?|CIRCUNSCRI|ATUACAO|POLICIAMENTO|ESTRUTURA|BAIRROS?|DESCRICAO)\b",
    flags=re.I,
)

RUIDO_ADMINISTRATIVO_RE = re.compile(
    r"\bMEDALHA\b|\bHOMENAGEAR\b|\bHOMENAGEM\b|"
    r"\bPROEIS\b|\bPROESP\b|\bREGIME\s+ADICIONAL\s+DE\s+SERVICO\b|\bRPE\b|"
    r"\bCONVENIOS?\b|\bTERMOS?\s+DE\s+COOPERACAO\b|\bINSTRUMENTOS?\s+CONGENERES\b|"
    r"\bREGIMENTO\s+INTERNO\b|\bAPROVA\s+O\s+REGIMENTO\b|"
    r"\bCOMISSAO\s+DE\s+FISCALIZACAO\b|\bGESTORES?\s+ESTRATEGIC[OA]S?\b|"
    r"\bFISCAIS?\s+DE\s+CONTRATO\b|\bCONTRATOS?\s+ADMINISTRATIVOS?\b",
    flags=re.I,
)

INDICADOR_FORTE_AREA_RE = re.compile(
    r"\b(?:AISP|RISP|CISP)\b|"
    r"\b(?:CIRCUNSCRICAO|CIRCUNSCRICOES|CIRCUNSCRICIONAL|CIRCUNSCRICIONAIS)\b|"
    r"\bLIMITES?\s+(?:TERRITORIAIS|CIRCUNSCRICIONAIS)\b|"
    r"\bDELIMITACAO\s+TERRITORIAL\b|"
    r"\b(?:BAIRRO|BAIRROS|MUNICIPIO|MUNICIPIOS)\b.{0,180}\b(?:DP|DELEGACIA|BPM|CISP|AISP|RISP)\b|"
    r"\b(?:DP|DELEGACIA|BPM)\b.{0,180}\b(?:BAIRROS?|MUNICIPIOS?|DESCRICAO|LIMITES?|CIRCUNSCRI|AREA\s+DE\s+(?:ATUACAO|POLICIAMENTO|RESPONSABILIDADE)|ESTRUTURA)\b|"
    r"\b(?:AREA|AREAS)\s+DE\s+(?:ATUACAO|POLICIAMENTO|RESPONSABILIDADE)\b.{0,180}\b(?:DP|DELEGACIA|BPM|AISP|CISP|RISP)\b|"
    r"\b(?:CRIA|CRIAR|CRIADO|CRIADA|FICA\s+CRIAD[AO]|EXTINGUE|EXTINGUIR|EXTINT[AO]|ALTERA|ALTERAR|MODIFICA|RESTRINGE|TRANSFERE|TRANSFERIR|TRANSFERINDO|PASSA\s+A\s+INTEGRAR|PASSANDO\s+A)\b.{0,240}\b(?:DP|DELEGACIA|BPM|CISP|AISP|RISP|COMPANHIA\s+INDEPENDENTE|BATALHAO)\b",
    flags=re.I,
)

INDICADOR_TERRITORIAL_EXPLICITO_RE = re.compile(
    r"\b(?:AISP|RISP|CISP)\b|"
    r"\bCIRCUNSCRI(?:CAO|COES|CIONAL|CIONAIS)\b|"
    r"\bLIMITES?\s+(?:TERRITORIAIS|CIRCUNSCRICIONAIS)\b|"
    r"\b(?:BAIRRO|BAIRROS|MUNICIPIO|MUNICIPIOS|DELIMITACAO\s+TERRITORIAL)\b",
    flags=re.I,
)

MUDANCA_UNIDADE_AREA_RE = re.compile(
    r"\b(?:CRIA|CRIAR|CRIADO|CRIADA|FICA\s+CRIAD[AO]|EXTINGUE|EXTINGUIR|EXTINT[AO]|ALTERA|ALTERAR|MODIFICA|TRANSFERE|TRANSFERIR|TRANSFERINDO|PASSA\s+A\s+INTEGRAR|PASSANDO\s+A)\b"
    r".{0,240}\b(?:DP|DELEGACIA|BPM|CISP|AISP|RISP|COMPANHIA\s+INDEPENDENTE|BATALHAO)\b",
    flags=re.I,
)

SIGLAS_ESCOPO_SEGURANCA = {
    "SESEG",
    "SSP",
    "SESP",
    "SEPOL",
    "SEPM",
    "PCERJ",
    "PMERJ",
}

TERMOS_ESCOPO_SEM_SIGLA = [
    "SEGURANCA PUBLICA",
    "SECRETARIA DE ESTADO DE POLICIA CIVIL",
    "SECRETARIA DE ESTADO DE POLICIA MILITAR",
    "POLICIA CIVIL",
    "POLICIA MILITAR",
    "INSTITUTO DE SEGURANCA PUBLICA",
    "ISP/RJ",
    "ISP-RJ",
]

PUBLICATION_ID_RE = re.compile(r"(?i)\bId\s*:?\s*\d+\.?")
RIO_DE_JANEIRO_RE = re.compile(r"(?i)\bRio\s+de\s+Janeiro\s*,")
NEXT_PUBLICATION_HEADING_AFTER_SIGNATURE_RE = re.compile(
    r"(?i)(?:^|\s)#{1,6}\s*("
    r"SECRETARIA\s+DE\s+ESTADO|"
    r"POL[ÍI]CIA\s+CIVIL|"
    r"CORREGEDORIA|"
    r"ATO\s+DO|"
    r"ATOS\s+DO|"
    r"DESPACHO\s+DO|"
    r"DESPACHOS\s+DO"
    r")\b"
)
MARCADORES_VAZAMENTO_RE = re.compile(
    r"##\s*(?:CORREGEDORIA|POL[ÍI]CIA\s+CIVIL|SECRETARIA)|"
    r"\|\s*Homic[íi]dio\s+Doloso\s*\|",
    flags=re.I,
)


def montar_texto_busca(row: pd.Series) -> str:
    campos = [
        row.get("secretaria_bloco", ""),
        row.get("autoridade_bloco", ""),
        row.get("tipo_norma", ""),
        row.get("sigla_norma", ""),
        row.get("ementa", ""),
        row.get("considerando", ""),
        row.get("mudancas_resumo", ""),
        row.get("assinatura", ""),
        row.get("texto_norma", ""),
    ]
    return normalizar_texto(" ".join(str(campo or "") for campo in campos))


def contem_circunscricao(texto: str) -> bool:
    return bool(CIRCUNSCRICAO_RE.search(texto))


def contem_area_atuacao(texto: str) -> bool:
    return bool(CIRCUNSCRICAO_RE.search(texto) or AREA_ATUACAO_RE.search(texto))


def texto_tem_indicio_forte_area(texto: str) -> bool:
    return bool(INDICADOR_FORTE_AREA_RE.search(texto))


def texto_eh_ruido_administrativo(texto: str) -> bool:
    if not RUIDO_ADMINISTRATIVO_RE.search(texto):
        return False

    return not (
        INDICADOR_TERRITORIAL_EXPLICITO_RE.search(texto)
        or MUDANCA_UNIDADE_AREA_RE.search(texto)
    )


def linha_tem_relevancia_area(row: pd.Series) -> bool:
    texto = row.get("texto_busca", "")

    if not texto_tem_indicio_forte_area(texto):
        return False

    return not texto_eh_ruido_administrativo(texto)


def recortar_texto_norma_para_artefato(value: object) -> str:
    texto = str(value or "").strip()

    if not texto:
        return ""

    candidatos = []

    id_match = PUBLICATION_ID_RE.search(texto)
    if id_match:
        candidatos.append(id_match.end())

    assinatura_match = RIO_DE_JANEIRO_RE.search(texto)
    if assinatura_match:
        next_heading = NEXT_PUBLICATION_HEADING_AFTER_SIGNATURE_RE.search(
            texto,
            assinatura_match.end(),
        )
        if next_heading:
            candidatos.append(next_heading.start())

    if not candidatos:
        return texto

    return texto[: min(candidatos)].strip()


def texto_tem_escopo_sem_sigla(*partes: object) -> bool:
    texto = normalizar_texto(" ".join(str(parte or "") for parte in partes))
    return any(termo in texto for termo in TERMOS_ESCOPO_SEM_SIGLA)


def linha_no_escopo_seguranca(row: pd.Series) -> bool:
    sigla = normalizar_texto(row.get("sigla_norma", ""))

    if sigla in SIGLAS_ESCOPO_SEGURANCA:
        return True

    if sigla:
        return False

    return texto_tem_escopo_sem_sigla(
        row.get("ementa", ""),
        row.get("considerando", ""),
        row.get("mudancas_resumo", ""),
        str(row.get("texto_norma", ""))[:5000],
    )


def limpar_ementa_para_artefato(row: pd.Series) -> str:
    ementa = limpar_texto(row.get("ementa", ""))

    if ementa and len(ementa) <= 3000 and not MARCADORES_VAZAMENTO_RE.search(ementa):
        return ementa

    texto = limpar_texto(row.get("texto_norma", ""))

    if not texto:
        return "" if len(ementa) > 3000 else ementa

    header = (
        r"^(?:RESOLU[ÇC][ÃA]O|PORTARIA|DECRETO|ATO|INSTRU[ÇC][ÃA]O\s+NORMATIVA)"
        r"(?:\s+[A-Z]{2,15})?\s+(?:N[º°O]\.?|N\.º|Nº|NO|N\.)\s*[\d./-]+"
        r"\s+DE\s+[^.]+\.?\s*"
    )
    trecho = re.sub(header, "", texto, flags=re.I)

    stops = [
        r"\bO\s+SECRET[ÁA]RIO\b",
        r"\bO\s+GOVERNADOR\b",
        r"\bCONSIDERANDO\b",
        r"\bRESOLVE\b",
        r"\bArt\.?\s*1",
    ]

    indices = []
    for pattern in stops:
        match = re.search(pattern, trecho, flags=re.I)
        if match:
            indices.append(match.start())

    if indices:
        trecho = trecho[: min(indices)]

    trecho = limpar_texto(trecho)
    return trecho[:1500].strip()


def limpar_dataframe_entrada(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    df = df.copy()
    df["texto_norma"] = df["texto_norma"].apply(recortar_texto_norma_para_artefato)
    df["ementa"] = df.apply(limpar_ementa_para_artefato, axis=1)

    mascara_escopo = df.apply(linha_no_escopo_seguranca, axis=1)
    df_limpo = df[mascara_escopo].copy()
    df_descartado = df[~mascara_escopo].copy()

    return df_limpo, df_descartado


def extrair_trechos_circunscricao(texto_original: object) -> str:
    texto = limpar_texto(texto_original)

    if not texto:
        return ""

    partes = re.split(r"(?<=[.;:])\s+|\s+\|\s+", texto)
    trechos = []
    vistos = set()

    for parte in partes:
        parte_limpa = limpar_texto(parte)

        if not parte_limpa:
            continue

        parte_norm = normalizar_texto(parte_limpa)

        if texto_tem_indicio_forte_area(parte_norm) and not texto_eh_ruido_administrativo(parte_norm):
            chave = parte_norm

            if chave not in vistos:
                vistos.add(chave)
                trechos.append(parte_limpa)

    return " | ".join(trechos)


def detectar_orgao_area(texto: str) -> str:
    regras = [
        ("Polícia Civil / PCERJ", [
            "POLICIA CIVIL", "POLÍCIA CIVIL", "PCERJ",
            "DELEGACIA DE POLICIA", "DELEGACIA POLICIAL",
            "DELEGACIAS DE POLICIA", "DELEGADO DE POLICIA",
        ]),
        ("Polícia Militar / PMERJ", [
            "POLICIA MILITAR", "POLÍCIA MILITAR", "PMERJ",
            "BATALHAO DE POLICIA MILITAR", "BATALHÃO DE POLÍCIA MILITAR",
        ]),
        ("UPP / Polícia Pacificadora", [
            "UNIDADE DE POLICIA PACIFICADORA",
            "UNIDADE DE POLÍCIA PACIFICADORA",
            " UPP", "UPP/",
        ]),
        ("SESEG / Secretaria de Estado de Segurança", [
            "SECRETARIA DE ESTADO DE SEGURANCA",
            "SECRETARIA DE ESTADO DE SEGURANÇA",
            "SECRETARIA DE SEGURANCA",
            "SECRETARIA DE SEGURANÇA",
            "SESEG",
        ]),
        ("ISP / Instituto de Segurança Pública", [
            "INSTITUTO DE SEGURANCA PUBLICA",
            "INSTITUTO DE SEGURANÇA PÚBLICA",
            " ISP ", "ISP/RJ", "ISP-RJ",
        ]),
    ]

    for area, keywords in regras:
        if any(normalizar_texto(keyword) in texto for keyword in keywords):
            return area

    return "Não identificado"


def classificar_tipo_circunscricao(texto: str) -> str:
    tipos = []

    if re.search(r"\bRESTAURANDO\b|\bRESTAURA\b|\bRESTABELEC", texto):
        tipos.append("restauração de limites/circunscrição")

    if re.search(r"\bALTERAR\b|\bALTERA\b|\bALTERACAO\b|\bALTERA[ÇC][AÃ]O\b|\bMODIFICA\b|\bMODIFICAD[AO]\b|\bREESTRUTURA", texto):
        tipos.append("alteração de circunscrição")

    if re.search(r"\bREVOGAD[AO]\b|\bREVOGAR\b|\bREVOGA\b", texto):
        tipos.append("revogação relacionada à circunscrição")

    if re.search(r"\bEXTINGUIR\b|\bEXTINT[AO]\b|\bEXTINCAO\b|\bEXTIN[ÇC][AÃ]O\b", texto):
        tipos.append("extinção com efeito territorial/circunscricional")

    if re.search(r"\bCRIAR\b|\bCRIAD[AO]\b|\bINSTITUIR\b|\bINSTITUI\b|\bIMPLANTA\b|\bIMPLANTAR\b", texto):
        tipos.append("criação/instituição com efeito circunscricional")

    if re.search(r"\bLIMITES?\b|\bAREA\b|\b[ÁA]REA\b|\bTERRITORIO\b|\bTERRIT[ÓO]RIO\b", texto):
        tipos.append("definição de limites/área territorial")

    if not tipos:
        return "menção circunscricional não classificada"

    saida = []
    vistos = set()

    for item in tipos:
        if item not in vistos:
            vistos.add(item)
            saida.append(item)

    return "; ".join(saida)


def detectar_tipo_mudanca(texto: str) -> str:
    regras = [
        ("revogação", r"\bREVOGAR\b|\bREVOGA\b|\bREVOGANDO\b|\bREVOGAD[AO]\b"),
        ("alteração", r"\bALTERAR\b|\bALTERAD[AO]\b|\bALTERA[ÇC][AÃ]O\b|\bMODIFICA\b|\bMODIFICAD[AO]\b|\bREESTRUTURA"),
        ("alteração", r"\bPASSA\s+A\s+(?:INTEGRAR|VIGORAR)\b|\bPASSANDO\s+A\b|\bTRANSFERINDO\b|\bTRANSFERIR\b"),
        ("extinção", r"\bEXTINGUIR\b|\bEXTINT[AO]\b|\bEXTINCAO\b|\bEXTIN[ÇC][AÃ]O\b"),
        ("criação/instituição", r"\bCRIAR\b|\bCRIAD[AO]\b|\bINSTITUIR\b|\bINSTITUI\b|\bIMPLANTA\b|\bIMPLANTAR\b"),
        ("competência/responsabilidade", r"\bCABERA\b|\bCABER[ÁA]\b|\bCOMPETE\b|\bDEVERA\b|\bDEVER[ÁA]\b|\bATRIBUI[ÇC][ÕO]ES\b|\bSER[ÁA]\s+DIRIGID[AO]\b"),
        ("vigência", r"\bENTRA EM VIGOR\b|\bENTRARA EM VIGOR\b|\bENTRAR[ÁA] EM VIGOR\b"),
    ]

    encontrados = [
        label
        for label, pattern in regras
        if re.search(pattern, texto, flags=re.I)
    ]

    return "; ".join(encontrados) if encontrados else "não classificado"


def parse_artigos_json(value: object) -> list[dict]:
    if pd.isna(value) or not str(value).strip():
        return []

    try:
        parsed = json.loads(str(value))
    except json.JSONDecodeError:
        return []

    if isinstance(parsed, list):
        return [item for item in parsed if isinstance(item, dict)]

    return []


def artigo_eh_relevante(artigo: dict) -> bool:
    texto = normalizar_texto(artigo.get("texto", ""))
    acao = normalizar_texto(artigo.get("acao", ""))

    if acao == "VIGENCIA":
        return False

    if texto_tem_indicio_forte_area(texto):
        return not texto_eh_ruido_administrativo(texto)

    if texto_eh_ruido_administrativo(texto):
        return False

    termos_territoriais = [
        "LIMITES", "AREA", "ÁREA", "TERRITORIO", "TERRITÓRIO",
        "DELEGACIA", "DELEGACIAS", "UPP", "BATALHAO",
        "BATALHÃO", "POLICIAMENTO",
    ]

    return bool(INDICADOR_TERRITORIAL_EXPLICITO_RE.search(texto)) and any(
        normalizar_texto(t) in texto for t in termos_territoriais
    )


def montar_o_que_mudou(row: pd.Series) -> str:
    artigos = parse_artigos_json(row.get("artigos_json", ""))

    partes = []

    for artigo in artigos:
        if not artigo_eh_relevante(artigo):
            continue

        label = limpar_texto(artigo.get("artigo", ""))
        acao = limpar_texto(artigo.get("acao", ""))
        objeto = limpar_texto(artigo.get("objeto", "")) or limpar_texto(artigo.get("texto", ""))

        if not objeto:
            continue

        if acao and acao.lower() not in {"outros", "não classificado", "nao classificado"}:
            partes.append(f"{label}: {acao} — {objeto}")
        else:
            partes.append(f"{label}: {objeto}")

    if partes:
        return " | ".join(partes)

    return limpar_texto(row.get("mudancas_resumo", ""))


def extrair_normas_referenciadas(texto_original: object) -> str:
    texto = limpar_texto(texto_original)

    patterns = [
        r"\bLei\s+n[º°]?\s*[\d./-]+(?:,\s*de\s*[^.;,\n]+)?",
        r"\bDecreto(?:-Lei)?\s+n[º°]?\s*[\d./-]+(?:,\s*de\s*[^.;,\n]+)?",
        r"\bResolu[çc][ãa]o\s+[A-Z]{0,12}\s*n[º°]?\s*[\d./-]+(?:,\s*de\s*[^.;,\n]+)?",
        r"\bPortaria\s+[A-Z]{0,12}\s*n[º°]?\s*[\d./-]+(?:,\s*de\s*[^.;,\n]+)?",
    ]

    encontrados = []

    for pattern in patterns:
        encontrados.extend(re.findall(pattern, texto, flags=re.I))

    vistos = set()
    saida = []

    for item in encontrados:
        item_limpo = limpar_texto(item)
        chave = normalizar_texto(item_limpo)

        if chave in vistos:
            continue

        vistos.add(chave)
        saida.append(item_limpo)

    return " | ".join(saida)


def montar_norma_id(row: pd.Series) -> str:
    tipo = limpar_texto(row.get("tipo_norma", ""))
    sigla = limpar_texto(row.get("sigla_norma", ""))
    numero = limpar_texto(row.get("numero_norma", ""))
    ano = str(row.get("ano_publicacao", "") or "")

    partes = [tipo]

    if sigla:
        partes.append(sigla)

    if numero:
        partes.append(f"nº {numero}")

    if ano:
        partes.append(f"({ano})")

    return " ".join(partes).strip()


def tratar_normativas(input_path: Path) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    colunas_esperadas = [
        "estado", "diario", "data_publicacao", "arquivo_markdown", "link",
        "secretaria_bloco", "autoridade_bloco", "tipo_norma", "sigla_norma",
        "numero_norma", "data_norma_texto", "ementa", "considerando",
        "mudancas_resumo", "artigos_json", "assinatura", "texto_norma",
    ]

    df = pd.read_csv(input_path, dtype=str, keep_default_na=False)
    df = garantir_colunas(df, colunas_esperadas)
    df = df.fillna("")
    linhas_entrada_bruta = len(df)

    if "id_linha_original" not in df.columns:
        df.insert(0, "id_linha_original", range(1, len(df) + 1))

    df, df_descartado_limpeza = limpar_dataframe_entrada(df)

    df["data_publicacao_dt"] = pd.to_datetime(df["data_publicacao"], errors="coerce")
    df["ano_publicacao"] = df["data_publicacao_dt"].dt.year.astype("Int64").astype(str)
    df["ano_publicacao"] = df["ano_publicacao"].replace("<NA>", "")

    df["texto_busca"] = df.apply(montar_texto_busca, axis=1)
    df["manter_circunscricao"] = df.apply(linha_tem_relevancia_area, axis=1)

    df["ementa_limpa"] = df["ementa"].apply(limpar_texto)
    df["orgao_area_detectada"] = df["texto_busca"].apply(detectar_orgao_area)
    df["tipo_circunscricao"] = df["texto_busca"].apply(classificar_tipo_circunscricao)
    df["tipo_mudanca_detectada"] = df["texto_busca"].apply(detectar_tipo_mudanca)
    df["qtd_artigos_extraidos"] = df["artigos_json"].apply(lambda x: len(parse_artigos_json(x)))
    df["o_que_mudou_tratado"] = df.apply(montar_o_que_mudou, axis=1)
    df["trechos_circunscricao"] = df["texto_norma"].apply(extrair_trechos_circunscricao)
    df["normas_referenciadas"] = df["texto_norma"].apply(extrair_normas_referenciadas)
    df["norma_id"] = df.apply(montar_norma_id, axis=1)

    df_tratado = df[df["manter_circunscricao"]].copy()

    colunas_prioritarias = [
        "ano_publicacao",
        "data_publicacao",
        "link",
        "norma_id",
        "tipo_norma",
        "sigla_norma",
        "numero_norma",
        "data_norma_texto",
        "orgao_area_detectada",
        "tipo_circunscricao",
        "tipo_mudanca_detectada",
        "ementa_limpa",
        "trechos_circunscricao",
        "o_que_mudou_tratado",
        "normas_referenciadas",
        "assinatura",
        "arquivo_markdown",
    ]

    outras_colunas = [col for col in df_tratado.columns if col not in colunas_prioritarias]
    df_tratado = df_tratado[colunas_prioritarias + outras_colunas].copy()

    df_mudancas = df_tratado[
        df_tratado["tipo_mudanca_detectada"].ne("não classificado")
    ].copy()

    df_mudancas = df_mudancas.sort_values(
        by=["data_publicacao", "tipo_norma", "numero_norma"],
        ascending=[True, True, True],
    )

    df_resumo = montar_resumo(
        df,
        df_tratado,
        df_mudancas,
        linhas_entrada_bruta=linhas_entrada_bruta,
        df_descartado_limpeza=df_descartado_limpeza,
    )

    return df_tratado, df_mudancas, df_resumo


def montar_resumo(
    df_original: pd.DataFrame,
    df_tratado: pd.DataFrame,
    df_mudancas: pd.DataFrame,
    linhas_entrada_bruta: int | None = None,
    df_descartado_limpeza: pd.DataFrame | None = None,
) -> pd.DataFrame:
    linhas_descartadas_limpeza = 0 if df_descartado_limpeza is None else len(df_descartado_limpeza)

    linhas = [
        ("linhas_entrada_bruta", len(df_original) if linhas_entrada_bruta is None else linhas_entrada_bruta),
        ("linhas_descartadas_limpeza_escopo", linhas_descartadas_limpeza),
        ("linhas_entrada", len(df_original)),
        ("linhas_com_circunscricao", len(df_tratado)),
        ("linhas_sem_circunscricao_descartadas", int((~df_original["manter_circunscricao"]).sum())),
        ("mudancas_circunscricao_classificadas", len(df_mudancas)),
        ("linhas_com_link", int(df_tratado["link"].astype(str).str.strip().ne("").sum())),
    ]

    for tipo, qtd in df_tratado["tipo_circunscricao"].value_counts().items():
        linhas.append((f"tipo_circunscricao: {tipo}", int(qtd)))

    for area, qtd in df_tratado["orgao_area_detectada"].value_counts().items():
        linhas.append((f"orgao_area_detectada: {area}", int(qtd)))

    for ano, qtd in df_tratado["ano_publicacao"].value_counts().sort_index().items():
        linhas.append((f"ano_publicacao: {ano}", int(qtd)))

    for tipo, qtd in df_tratado["tipo_norma"].value_counts().items():
        linhas.append((f"tipo_norma: {tipo}", int(qtd)))

    return pd.DataFrame(linhas, columns=["indicador", "valor"])


def selecionar_colunas_xlsx(df: pd.DataFrame) -> pd.DataFrame:
    colunas_xlsx = [
        "ano_publicacao",
        "data_publicacao",
        "link",
        "norma_id",
        "orgao_area_detectada",
        "tipo_circunscricao",
        "tipo_mudanca_detectada",
        "ementa_limpa",
        "trechos_circunscricao",
        "o_que_mudou_tratado",
        "normas_referenciadas",
        "assinatura",
        "arquivo_markdown",
    ]

    colunas_existentes = [col for col in colunas_xlsx if col in df.columns]
    return df[colunas_existentes].copy()


def ajustar_excel(xlsx_path: Path) -> None:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = load_workbook(xlsx_path)

    cores = {
        "Resumo": "1F4E78",
        "Base_circunscricao": "305496",
        "Mudancas": "548235",
        "Dicionario": "7030A0",
    }

    larguras_padrao = {
        "ano_publicacao": 12,
        "data_publicacao": 14,
        "link": 45,
        "norma_id": 28,
        "orgao_area_detectada": 30,
        "tipo_circunscricao": 38,
        "tipo_mudanca_detectada": 32,
        "ementa_limpa": 55,
        "trechos_circunscricao": 65,
        "o_que_mudou_tratado": 70,
        "normas_referenciadas": 45,
        "assinatura": 35,
        "arquivo_markdown": 42,
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        cor = cores.get(sheet_name, "1F4E78")

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=cor)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        header_by_col = {}
        for idx, cell in enumerate(ws[1], start=1):
            header_by_col[idx] = str(cell.value or "")

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

                if header_by_col.get(cell.column) == "link" and cell.row > 1 and cell.value:
                    cell.hyperlink = str(cell.value)
                    cell.style = "Hyperlink"

        for idx, header in header_by_col.items():
            column_letter = get_column_letter(idx)
            ws.column_dimensions[column_letter].width = larguras_padrao.get(header, 18)

        for row_idx in range(2, min(ws.max_row, 300) + 1):
            ws.row_dimensions[row_idx].height = 60

    wb.save(xlsx_path)


def exportar_resultados(
    df_tratado: pd.DataFrame,
    df_mudancas: pd.DataFrame,
    df_resumo: pd.DataFrame,
) -> None:
    PASTA_SAIDA.mkdir(parents=True, exist_ok=True)

    tratado_csv = PASTA_SAIDA / "ISP_circunscricao_tratado.csv"
    mudancas_csv = PASTA_SAIDA / "ISP_circunscricao_mudancas.csv"
    resumo_csv = PASTA_SAIDA / "ISP_circunscricao_resumo.csv"
    xlsx_path = PASTA_SAIDA / "ISP_circunscricao_tratado.xlsx"

    df_tratado.to_csv(tratado_csv, index=False, encoding="utf-8-sig")
    df_mudancas.to_csv(mudancas_csv, index=False, encoding="utf-8-sig")
    df_resumo.to_csv(resumo_csv, index=False, encoding="utf-8-sig")

    if GERAR_EXCEL:
        dicionario = pd.DataFrame(
            [
                ("link", "URL da página da edição de onde o Markdown foi gerado, recuperada pelo extrator com edition.url."),
                ("tipo_circunscricao", "Classifica se houve alteração, restauração, revogação, definição de limites etc."),
                ("trechos_circunscricao", "Frases/trechos em que aparecem os termos de circunscrição."),
                ("o_que_mudou_tratado", "Resumo dos artigos relevantes para a mudança territorial/circunscricional."),
                ("normas_referenciadas", "Normas citadas no texto, quando detectadas por regex."),
                ("arquivo_markdown", "Arquivo local extraído pelo Docling."),
            ],
            columns=["campo", "descricao"],
        )

        df_tratado_xlsx = selecionar_colunas_xlsx(df_tratado)
        df_mudancas_xlsx = selecionar_colunas_xlsx(df_mudancas)

        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
            df_resumo.to_excel(writer, sheet_name="Resumo", index=False)
            df_tratado_xlsx.to_excel(writer, sheet_name="Base_circunscricao", index=False)
            df_mudancas_xlsx.to_excel(writer, sheet_name="Mudancas", index=False)
            dicionario.to_excel(writer, sheet_name="Dicionario", index=False)

        ajustar_excel(xlsx_path)

    print("\nArquivos gerados:")
    print(f"- {tratado_csv}")
    print(f"- {mudancas_csv}")
    print(f"- {resumo_csv}")

    if GERAR_EXCEL:
        print(f"- {xlsx_path}")


def main() -> None:
    print("=" * 70)
    print("TRATAMENTO — CIRCUNSCRIÇÃO / ISP COM COLUNA link")
    print("=" * 70)

    print(f"\nPasta base: {BASE_DIR}")
    print(f"Arquivo de entrada esperado: {ARQUIVO_ENTRADA}")
    print(f"Pasta de saída: {PASTA_SAIDA}")

    if not ARQUIVO_ENTRADA.exists():
        print("\nERRO: não encontrei o arquivo de entrada.")
        print("Rode primeiro o extrator com link ou confira o caminho:")
        print(ARQUIVO_ENTRADA)
        raise SystemExit(1)

    print("\nLendo e tratando o CSV de circunscrição...")
    df_tratado, df_mudancas, df_resumo = tratar_normativas(ARQUIVO_ENTRADA)

    print("Exportando resultados...")
    exportar_resultados(
        df_tratado=df_tratado,
        df_mudancas=df_mudancas,
        df_resumo=df_resumo,
    )

    print("\nResumo:")
    print(df_resumo.to_string(index=False))

    print("\nProcesso concluído com sucesso.")


if __name__ == "__main__":
    main()
